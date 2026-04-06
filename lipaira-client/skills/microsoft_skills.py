"""
Microsoft Graph skills covering Outlook, Calendar,
OneDrive, Word, Excel, and OneNote.
"""
import os
import base64
from datetime import datetime, timedelta
from .base import BaseSkill, SkillResult

from .base import BaseSkill, SkillResult
from .microsoft_client import graph_get, graph_post, graph_patch, graph_put_bytes


# ── Outlook: Send ────────────────────────────────────────────────────────────

class OutlookSendSkill(BaseSkill):
    name = "outlook_send"
    description = (
        "Send an email from the user's Outlook/Microsoft account. "
        "Use for sending invoices, quotes, follow-ups, and client communications."
    )

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "to": {"type": "string"},
                "subject": {"type": "string"},
                "body": {"type": "string"},
                "cc": {"type": "string"}
            },
            "required": ["to", "subject", "body"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            message = {
                "subject": input["subject"],
                "body": {
                    "contentType": "HTML" if "<" in input["body"] else "Text",
                    "content": input["body"]
                },
                "toRecipients": [
                    {"emailAddress": {"address": input["to"]}}
                ]
            }

            if input.get("cc"):
                message["ccRecipients"] = [
                    {"emailAddress": {"address": input["cc"]}}
                ]

            graph_post("/me/sendMail", {"message": message})

            return SkillResult(success=True, output={
                "message": f"Email sent to {input['to']}",
                "subject": input["subject"]
            })
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── Outlook: Read ────────────────────────────────────────────────────────────

class OutlookReadSkill(BaseSkill):
    name = "outlook_read"
    description = (
        "Read recent emails from the user's Outlook inbox. "
        "Use to check for client replies or urgent messages."
    )

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "max_results": {"type": "integer", "default": 5},
                "unread_only": {"type": "boolean", "default": False}
            }
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            max_results = min(input.get("max_results", 5), 20)
            params = {
                "$top": max_results,
                "$select": "id,subject,from,receivedDateTime,bodyPreview,isRead",
                "$orderby": "receivedDateTime desc"
            }

            data = graph_get("/me/messages", params)

            emails = []
            for msg in data.get("value", []):
                emails.append({
                    "subject": msg.get("subject", ""),
                    "from": msg.get("from", {}).get("emailAddress", {}).get("address", ""),
                    "date": msg.get("receivedDateTime", ""),
                    "preview": msg.get("bodyPreview", ""),
                    "unread": not msg.get("isRead", True)
                })

            return SkillResult(success=True, output=emails)
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── Outlook Calendar: Read ───────────────────────────────────────────────────

class OutlookCalendarReadSkill(BaseSkill):
    name = "outlook_calendar_read"
    description = (
        "Read upcoming events from the user's Outlook Calendar. "
        "Use to check availability or list appointments."
    )

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "days_ahead": {"type": "integer", "default": 7},
                "max_results": {"type": "integer", "default": 10}
            }
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            now = datetime.utcnow()
            end = now + timedelta(days=input.get("days_ahead", 7))

            data = graph_get("/me/calendarView", {
                "startDateTime": now.isoformat() + "Z",
                "endDateTime": end.isoformat() + "Z",
                "$top": min(input.get("max_results", 10), 25),
                "$select": "id,subject,start,end,location,attendees"
            })

            events = []
            for e in data.get("value", []):
                events.append({
                    "title": e.get("subject", "Untitled"),
                    "start": e.get("start", {}).get("dateTime", ""),
                    "end": e.get("end", {}).get("dateTime", ""),
                    "location": e.get("location", {}).get("displayName", "")
                })

            return SkillResult(success=True, output={"events": events, "count": len(events)})
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── Outlook Calendar: Write ──────────────────────────────────────────────────

class OutlookCalendarWriteSkill(BaseSkill):
    name = "outlook_calendar_write"
    description = (
        "Create a calendar event in the user's Outlook Calendar. "
        "Sends meeting invites to attendees automatically."
    )

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "start": {"type": "string", "description": "ISO 8601"},
                "end": {"type": "string"},
                "description": {"type": "string"},
                "location": {"type": "string"},
                "attendees": {"type": "array", "items": {"type": "string"}},
                "add_teams": {"type": "boolean", "default": False}
            },
            "required": ["title", "start", "end"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            event = {
                "subject": input["title"],
                "body": {"contentType": "HTML", "content": input.get("description", "")},
                "start": {"dateTime": input["start"], "timeZone": "UTC"},
                "end": {"dateTime": input["end"], "timeZone": "UTC"},
                "location": {"displayName": input.get("location", "")}
            }

            if input.get("attendees"):
                event["attendees"] = [
                    {"emailAddress": {"address": email}, "type": "required"}
                    for email in input["attendees"]
                ]

            if input.get("add_teams"):
                event["isOnlineMeeting"] = True
                event["onlineMeetingProvider"] = "teamsForBusiness"

            created = graph_post("/me/events", event)

            return SkillResult(success=True, output={
                "event_id": created.get("id"),
                "title": created.get("subject"),
                "web_link": created.get("webLink"),
                "teams_link": created.get("onlineMeeting", {}).get("joinUrl", ""),
                "message": f"Event created: {created.get('webLink')}"
            })
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── OneDrive Upload ──────────────────────────────────────────────────────────

class OneDriveUploadSkill(BaseSkill):
    name = "onedrive_upload"
    description = "Upload a file from documents to OneDrive."

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "filename": {"type": "string"},
                "folder": {"type": "string", "default": "Lipaira Documents"}
            },
            "required": ["filename"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            local_path = f"/app/data/documents/{input['filename']}"
            if not os.path.exists(local_path):
                return SkillResult(success=False, output=None, error=f"File not found")

            folder = input.get("folder", "Lipaira Documents")
            filename = input["filename"]

            with open(local_path, "rb") as f:
                file_data = f.read()

            uploaded = graph_put_bytes(
                f"/me/drive/root:/{folder}/{filename}:/content",
                file_data, "application/pdf"
            )

            return SkillResult(success=True, output={
                "web_url": uploaded.get("webUrl"),
                "message": f"Uploaded to OneDrive"
            })
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── Word Document ────────────────────────────────────────────────────────────

class WordCreateSkill(BaseSkill):
    name = "word_create"
    description = "Create a Word document in OneDrive."

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "content": {"type": "string"},
                "folder": {"type": "string", "default": "Lipaira Documents"}
            },
            "required": ["title", "content"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            folder = input.get("folder", "Lipaira Documents")
            filename = f"{input['title'].replace(' ', '_')}.docx"
            content = input["content"].encode("utf-8")

            uploaded = graph_put_bytes(
                f"/me/drive/root:/{folder}/{filename}:/content",
                content,
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )

            return SkillResult(success=True, output={
                "web_url": uploaded.get("webUrl"),
                "message": f"Word document created"
            })
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── Excel Spreadsheet ────────────────────────────────────────────────────────

class ExcelCreateSkill(BaseSkill):
    name = "excel_create"
    description = "Create an Excel spreadsheet in OneDrive."

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "headers": {"type": "array", "items": {"type": "string"}},
                "rows": {"type": "array", "items": {"type": "array"}},
                "folder": {"type": "string", "default": "Lipaira Documents"}
            },
            "required": ["title", "headers"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            from openpyxl.styles import Font
            for col, header in enumerate(input["headers"], 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            for row_idx, row in enumerate(input.get("rows") or [], 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            folder = input.get("folder", "Lipaira Documents")
            filename = f"{input['title'].replace(' ', '_')}.xlsx"

            uploaded = graph_put_bytes(
                f"/me/drive/root:/{folder}/{filename}:/content",
                buffer.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            return SkillResult(success=True, output={
                "web_url": uploaded.get("webUrl"),
                "rows": len(input.get("rows") or []),
                "message": f"Excel spreadsheet created"
            })
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── OneNote ──────────────────────────────────────────────────────────────────

class OneNoteCreateSkill(BaseSkill):
    name = "onenote_create"
    description = "Create a OneNote page with content."

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "content": {"type": "string"},
                "section_name": {"type": "string", "default": "Lipaira"}
            },
            "required": ["title", "content"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            # Get access token directly
            from .microsoft_client import get_access_token
            token = get_access_token()

            section_name = input.get("section_name", "Lipaira")
            
            html = f"""
            <!DOCTYPE html>
            <html>
            <head><title>{input['title']}</title></head>
            <body>
            <h1>{input['title']}</h1>
            <p>{input['content'].replace(chr(10), '<br>')}</p>
            </body>
            </html>
            """

            import requests as req
            # Get sections
            sections = graph_get("/me/onenote/sections")
            section_items = sections.get("value", [])
            
            if section_items:
                section_id = section_items[0]["id"]
            else:
                # Get default notebook
                notebooks = graph_get("/me/onenote/notebooks")
                if notebooks.get("value"):
                    notebook_id = notebooks["value"][0]["id"]
                    new_section = graph_post(
                        f"/me/onenote/notebooks/{notebook_id}/sections",
                        {"displayName": section_name}
                    )
                    section_id = new_section["id"]
                else:
                    return SkillResult(success=False, output=None, error="No notebook found")

            # Create page
            resp = req.post(
                f"https://graph.microsoft.com/v1.0/me/onenote/sections/{section_id}/pages",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/xhtml+xml"},
                data=html.encode("utf-8"),
                timeout=30
            )
            resp.raise_for_status()
            page = resp.json()

            return SkillResult(success=True, output={
                "title": input["title"],
                "web_url": page.get("links", {}).get("oneNoteWebUrl", {}).get("href", ""),
                "message": f"OneNote page created"
            })
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))


# ── Outlook Contacts ─────────────────────────────────────────────────────────

class OutlookContactLookupSkill(BaseSkill):
    name = "outlook_contact_lookup"
    description = "Look up a contact from Outlook Contacts by name."

    def get_input_schema(self):
        return {
            "type": "object",
            "properties": {
                "name": {"type": "string"}
            },
            "required": ["name"]
        }

    def execute(self, input: dict) -> SkillResult:
        try:
            data = graph_get("/me/contacts", {
                "$filter": f"contains(displayName, '{input['name']}')",
                "$top": 5
            })

            contacts = []
            for c in data.get("value", []):
                emails = c.get("emailAddresses", [])
                phones = c.get("businessPhones", [])
                contacts.append({
                    "name": c.get("displayName", ""),
                    "email": emails[0].get("address", "") if emails else "",
                    "phone": phones[0] if phones else ""
                })

            return SkillResult(success=True, output=contacts)
        except RuntimeError as e:
            return SkillResult(success=False, output=None, error=str(e))
        except Exception as e:
            return SkillResult(success=False, output=None, error=str(e))